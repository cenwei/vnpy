""""""
from abc import ABC
from copy import copy
import datetime
import os
from os import path
import sys
import traceback
from typing import Any, Callable, Dict
import logging
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from vnpy.component.cta_grid_trade import CtaGrid, CtaGridTrade
from vnpy.component.cta_position import CtaPosition

from vnpy.trader.constant import Interval, Direction, Offset, OrderType, Status
from vnpy.trader.object import BarData, TickData, OrderData, TradeData
from vnpy.trader.utility import get_folder_path, virtual, append_data

from .base import StopOrder, EngineType


class CtaTemplate(ABC):
    """"""

    author = ""
    parameters = []
    variables = []

    def __init__(
        self,
        cta_engine: Any,
        strategy_name: str,
        vt_symbol: str,
        setting: dict,
    ):
        """"""
        self.cta_engine = cta_engine
        self.strategy_name = strategy_name
        self.vt_symbol = vt_symbol

        self.inited = False
        self.trading = False
        self.pos = 0
        self.entrust = 0  # 是否正在委托, 0, 无委托 , 1, 委托方向是LONG， -1, 委托方向是SHORT

        self.active_orders:Dict[str, Dict] = {}

        # Copy a new variables list here to avoid duplicate insert when multiple
        # strategy instances are created with the same strategy class.
        self.variables = copy(self.variables)
        self.variables.insert(0, "inited")
        self.variables.insert(1, "trading")
        self.variables.insert(2, "pos")
        self.variables.insert(3, "entrust")

        self.update_setting(setting)

    def update_setting(self, setting: dict):
        """
        Update strategy parameter wtih value in setting dict.
        """
        for name in self.parameters:
            if name in setting:
                setattr(self, name, setting[name])

    @classmethod
    def get_class_parameters(cls):
        """
        Get default parameters dict of strategy class.
        """
        class_parameters = {}
        for name in cls.parameters:
            class_parameters[name] = getattr(cls, name)
        return class_parameters

    def get_parameters(self):
        """
        Get strategy parameters dict.
        """
        strategy_parameters = {}
        for name in self.parameters:
            strategy_parameters[name] = getattr(self, name)
        return strategy_parameters

    def get_variables(self):
        """
        Get strategy variables dict.
        """
        strategy_variables = {}
        for name in self.variables:
            strategy_variables[name] = getattr(self, name)
        return strategy_variables

    def get_data(self):
        """
        Get strategy data.
        """
        strategy_data = {
            "strategy_name": self.strategy_name,
            "vt_symbol": self.vt_symbol,
            "class_name": self.__class__.__name__,
            "author": self.author,
            "parameters": self.get_parameters(),
            "variables": self.get_variables(),
        }
        return strategy_data

    @virtual
    def on_init(self):
        """
        Callback when strategy is inited.
        """
        pass

    @virtual
    def on_start(self):
        """
        Callback when strategy is started.
        """
        pass

    @virtual
    def on_stop(self):
        """
        Callback when strategy is stopped.
        """
        pass

    @virtual
    def on_tick(self, tick: TickData):
        """
        Callback of new tick data update.
        """
        pass

    @virtual
    def on_bar(self, bar: BarData):
        """
        Callback of new bar data update.
        """
        pass

    @virtual
    def on_trade(self, trade: TradeData):
        """
        Callback of new trade data update.
        """
        pass

    @virtual
    def on_order(self, order: OrderData):
        """
        Callback of new order data update.
        """
        pass

    @virtual
    def on_stop_order(self, stop_order: StopOrder):
        """
        Callback of stop order update.
        """
        pass

    def buy(self, price: float, volume: float, stop: bool = False, lock: bool = False,
            vt_symbol: str = '', order_type: OrderType = OrderType.LIMIT, order_time: datetime = None, grid: CtaGrid = None):
        """
        Send buy order to open a long position.
        """
        return self.send_order(vt_symbol, Direction.LONG, Offset.OPEN, price, volume, stop, lock, order_type=order_type, order_time=order_time, grid=grid)

    def sell(self, price: float, volume: float, stop: bool = False, lock: bool = False,
            vt_symbol: str = '', order_type: OrderType = OrderType.LIMIT, order_time: datetime = None, grid: CtaGrid = None):
        """
        Send sell order to close a long position.
        """
        return self.send_order(vt_symbol, Direction.SHORT, Offset.CLOSE, price, volume, stop, lock, order_type=order_type, order_time=order_time, grid=grid)

    def short(self, price: float, volume: float, stop: bool = False, lock: bool = False,
            vt_symbol: str = '', order_type: OrderType = OrderType.LIMIT, order_time: datetime = None, grid: CtaGrid = None):
        """
        Send short order to open as short position.
        """
        return self.send_order(vt_symbol, Direction.SHORT, Offset.OPEN, price, volume, stop, lock, order_type=order_type, order_time=order_time, grid=grid)

    def cover(self, price: float, volume: float, stop: bool = False, lock: bool = False,
            vt_symbol: str = '', order_type: OrderType = OrderType.LIMIT, order_time: datetime = None, grid: CtaGrid = None):
        """
        Send cover order to close a short position.
        """
        return self.send_order(vt_symbol, Direction.LONG, Offset.CLOSE, price, volume, stop, lock, order_type=order_type, order_time=order_time, grid=grid)

    def send_order(
        self,
        vt_symbol: str,
        direction: Direction,
        offset: Offset,
        price: float,
        volume: float,
        stop: bool = False,
        lock: bool = False,
        order_type: OrderType = OrderType.LIMIT,
        order_time: datetime = None,
        grid: CtaGrid = None
    ):
        """
        Send a new order.
        """

        if vt_symbol == '':
            vt_symbol = self.vt_symbol

        if self.trading:
            if direction == Direction.LONG:
                self.entrust = 1
            elif direction == Direction.SHORT:
                self.entrust = -1
            
            vt_orderids = self.cta_engine.send_order(
                self, direction, offset, price, volume, stop, False, order_type
            )
            for vt_orderid in vt_orderids:
                d = {
                    'direction': direction,
                    'offset': offset,
                    'vt_symbol': vt_symbol,
                    'price': price,
                    'volume': volume,
                    'order_type': order_type,
                    'traded': 0,
                    'order_time': order_time,
                    'status': Status.SUBMITTING
                }
                if grid:
                    d.update({'grid': grid})
                    if lock:
                        grid.lock_grid_ids.append(vt_orderid)
                    else:
                        grid.order_ids.append(vt_orderid)
                self.active_orders.update({vt_orderid: d})

            return vt_orderids
        else:
            return []

    def cancel_order(self, vt_orderid: str):
        """
        Cancel an existing order.
        """
        if self.trading:
            return self.cta_engine.cancel_order(self, vt_orderid)

    def cancel_all(self):
        """
        Cancel all orders sent by strategy.
        """
        if self.trading:
            self.entrust = 0
            self.cta_engine.cancel_all(self)

    def write_log(self, msg: str, level: int = logging.INFO):
        """
        Write a log message.
        """
        self.cta_engine.write_log(msg, self, level)

    def write_error(self, msg: str):
        """
        Write error log message
        """
        self.write_log(msg=msg, level=logging.ERROR)

    def get_engine_type(self):
        """
        Return whether the cta_engine is backtesting or live trading.
        """
        return self.cta_engine.get_engine_type()

    def get_pricetick(self):
        """
        Return pricetick data of trading contract.
        """
        return self.cta_engine.get_pricetick(self)

    def get_size(self):
        """
        返回杠杆倍数.
        """
        return self.cta_engine.get_size(self)

    def get_margin_rate(self):
        """
        返回保证金比率.
        """
        return self.cta_engine.get_margin_rate(self)

    def get_position_detail(self, vt_symbol: str):        
        """"""        
        return self.cta_engine.get_position_detail(vt_symbol)

    def load_bar(
        self,
        days: int,
        interval: Interval = Interval.MINUTE,
        callback: Callable = None,
        use_database: bool = False
    ):
        """
        Load historical bar data for initializing strategy.
        """
        if not callback:
            callback = self.on_bar

        self.cta_engine.load_bar(
            self.vt_symbol,
            days,
            interval,
            callback,
            use_database
        )

    def load_tick(self, days: int):
        """
        Load historical tick data for initializing strategy.
        """
        self.cta_engine.load_tick(self.vt_symbol, days, self.on_tick)

    def put_event(self):
        """
        Put an strategy data event for ui update.
        """
        if self.inited:
            self.cta_engine.put_strategy_event(self)

    def send_email(self, msg):
        """
        Send email to default receiver.
        """
        if self.inited:
            self.cta_engine.send_email(msg, self)

    def sync_data(self):
        """
        Sync strategy variables value into disk storage.
        """
        if self.trading:
            self.cta_engine.sync_strategy_data(self)


class CtaSignal(ABC):
    """"""

    def __init__(self):
        """"""
        self.signal_pos = 0

    @virtual
    def on_tick(self, tick: TickData):
        """
        Callback of new tick data update.
        """
        pass

    @virtual
    def on_bar(self, bar: BarData):
        """
        Callback of new bar data update.
        """
        pass

    def set_signal_pos(self, pos):
        """"""
        self.signal_pos = pos

    def get_signal_pos(self):
        """"""
        return self.signal_pos


class TargetPosTemplate(CtaTemplate):
    """"""
    tick_add = 1

    last_tick = None
    last_bar = None
    target_pos = 0

    def __init__(self, cta_engine, strategy_name, vt_symbol, setting):
        """"""
        super().__init__(cta_engine, strategy_name, vt_symbol, setting)

        self.active_orderids = []
        self.cancel_orderids = []

        self.variables.append("target_pos")

    @virtual
    def on_tick(self, tick: TickData):
        """
        Callback of new tick data update.
        """
        self.last_tick = tick

        if self.trading:
            self.trade()

    @virtual
    def on_bar(self, bar: BarData):
        """
        Callback of new bar data update.
        """
        self.last_bar = bar

    @virtual
    def on_order(self, order: OrderData):
        """
        Callback of new order data update.
        """
        vt_orderid = order.vt_orderid

        if not order.is_active():
            if vt_orderid in self.active_orderids:
                self.active_orderids.remove(vt_orderid)

            if vt_orderid in self.cancel_orderids:
                self.cancel_orderids.remove(vt_orderid)

    def check_order_finished(self):
        """"""
        if self.active_orderids:
            return False
        else:
            return True

    def set_target_pos(self, target_pos):
        """"""
        self.target_pos = target_pos
        self.trade()

    def trade(self):
        """"""
        if not self.check_order_finished():
            self.cancel_old_order()
        else:
            self.send_new_order()

    def cancel_old_order(self):
        """"""
        for vt_orderid in self.active_orderids:
            if vt_orderid not in self.cancel_orderids:
                self.cancel_order(vt_orderid)
                self.cancel_orderids.append(vt_orderid)

    def send_new_order(self):
        """"""
        pos_change = self.target_pos - self.pos
        if not pos_change:
            return

        long_price = 0
        short_price = 0

        if self.last_tick:
            if pos_change > 0:
                long_price = self.last_tick.ask_price_1 + self.tick_add
                if self.last_tick.limit_up:
                    long_price = min(long_price, self.last_tick.limit_up)
            else:
                short_price = self.last_tick.bid_price_1 - self.tick_add
                if self.last_tick.limit_down:
                    short_price = max(short_price, self.last_tick.limit_down)

        else:
            if pos_change > 0:
                long_price = self.last_bar.close_price + self.tick_add
            else:
                short_price = self.last_bar.close_price - self.tick_add

        if self.get_engine_type() == EngineType.BACKTESTING:
            if pos_change > 0:
                vt_orderids = self.buy(long_price, abs(pos_change))
            else:
                vt_orderids = self.short(short_price, abs(pos_change))
            self.active_orderids.extend(vt_orderids)

        else:
            if self.active_orderids:
                return

            if pos_change > 0:
                if self.pos < 0:
                    if pos_change < abs(self.pos):
                        vt_orderids = self.cover(long_price, pos_change)
                    else:
                        vt_orderids = self.cover(long_price, abs(self.pos))
                else:
                    vt_orderids = self.buy(long_price, abs(pos_change))
            else:
                if self.pos > 0:
                    if abs(pos_change) < self.pos:
                        vt_orderids = self.sell(short_price, abs(pos_change))
                    else:
                        vt_orderids = self.sell(short_price, abs(self.pos))
                else:
                    vt_orderids = self.short(short_price, abs(pos_change))
            self.active_orderids.extend(vt_orderids)

class CryptoFutureTemplate(CtaTemplate):
    """
    数字货币合约期货模板
    """
    size = 1.0

    def __init__(self, cta_engine, strategy_name, vt_symbol, setting):
        self.position:CtaPosition = None  # 仓位组件
        self.gt:CtaGridTrade = None  # 网格交易组件

        self.cur_datetime: datetime = None  # 当前Tick时间

        super().__init__(
            cta_engine, strategy_name, vt_symbol, setting
        )

        # 增加仓位管理模块
        self.position = CtaPosition(strategy=self)
        self.position.maxPos = sys.maxsize
        # 增加网格持久化模块
        self.gt = CtaGridTrade(strategy=self)

        self.size = self.get_size()

    def init_position(self):
        """
        初始化Positin
        使用网格的持久化，获取开仓状态的多空单，更新
        :return:
        """
        self.write_log(u'init_position(),初始化持仓')
        changed = False
        if len(self.gt.up_grids) <= 0:
            self.position.short_pos = 0
            # 加载已开仓的空单数据，网格JSON
            short_grids = self.gt.load(direction=Direction.SHORT, open_status_filter=[True])
            if len(short_grids) == 0:
                self.write_log(u'没有持久化的空单数据')
                self.gt.up_grids = []

            else:
                self.gt.up_grids = short_grids
                for sg in short_grids:

                    self.write_log(u'加载持仓空单[{},价格:{},数量:{}手,开仓时间:{}'
                                   .format(self.vt_symbol, sg.open_price,
                                           sg.volume, sg.open_time))
                    self.position.short_pos = round(self.position.short_pos - sg.volume, 7)

                self.write_log(u'持久化空单，共持仓:{}手'.format(abs(self.position.short_pos)))

        if len(self.gt.dn_grids) <= 0:
            # 加载已开仓的多数据，网格JSON
            self.position.long_pos = 0
            long_grids = self.gt.load(direction=Direction.LONG, open_status_filter=[True])
            if len(long_grids) == 0:
                self.write_log(u'没有持久化的多单数据')
                self.gt.dn_grids = []
            else:
                self.gt.dn_grids = long_grids
                for lg in long_grids:

                    self.write_log(u'加载持仓多单[{},价格:{},数量:{}手, 开仓时间:{}'
                                   .format(self.vt_symbol, lg.open_price, lg.volume, lg.open_time))
                    self.position.long_pos = round(self.position.long_pos + lg.volume, 7)

                self.write_log(f'持久化多单，共持仓:{self.position.long_pos}手')

        self.position.pos = round(self.position.long_pos + self.position.short_pos, 7)

        self.write_log(u'{}加载持久化数据完成，多单:{}，空单:{},共:{}手'
                       .format(self.strategy_name,
                               self.position.long_pos,
                               abs(self.position.short_pos),
                               self.position.pos))
        self.pos = self.position.pos
        if changed:
            self.gt.save()
        self.display_grids()

    def display_grids(self):
        """
        更新网格显示信息
        """
        if not self.inited:
            return
        self.account_pos = self.cta_engine.get_position(vt_symbol=self.vt_symbol, direction=Direction.NET)
        if self.account_pos:
            self.write_log(
                f'账号{self.vt_symbol}持仓:{self.account_pos.volume}, 冻结:{self.account_pos.frozen}, 盈亏:{self.account_pos.pnl}')

        up_grids_info = ""
        for grid in list(self.gt.up_grids):
            if not grid.open_status and grid.order_status:
                up_grids_info += f'平空中: [已平:{grid.traded_volume} => 目标:{grid.volume}, 委托时间:{grid.order_time}]\n'
                if len(grid.order_ids) > 0:
                    up_grids_info += f'委托单号:{grid.order_ids}'
                continue

            if grid.open_status and not grid.order_status:
                up_grids_info += f'持空中: [数量:{grid.volume}, 开仓时间:{grid.open_time}]\n'
                continue

            if not grid.open_status and grid.order_status:
                up_grids_info += f'开空中: [已开:{grid.traded_volume} => 目标:{grid.volume}, 委托时间:{grid.order_time}]\n'
                if len(grid.order_ids) > 0:
                    up_grids_info += f'委托单号:{grid.order_ids}'

        dn_grids_info = ""
        for grid in list(self.gt.dn_grids):
            if not grid.open_status and grid.order_status:
                dn_grids_info += f'平多中: [已平:{grid.traded_volume} => 目标:{grid.volume}, 委托时间:{grid.order_time}]\n'
                if len(grid.order_ids) > 0:
                    dn_grids_info += f'委托单号:{grid.order_ids}'
                continue

            if grid.open_status and not grid.order_status:
                dn_grids_info += f'持多中: [数量:{grid.volume}\n, 开仓时间:{grid.open_time}]\n'
                continue

            if not grid.open_status and grid.order_status:
                dn_grids_info += f'开多中: [已开:{grid.traded_volume} => 目标:{grid.volume}, 委托时间:{grid.order_time}]\n'
                if len(grid.order_ids) > 0:
                    dn_grids_info += f'委托单号:{grid.order_ids}'

        if len(up_grids_info) > 0:
            self.write_log(up_grids_info)
        if len(dn_grids_info) > 0:
            self.write_log(dn_grids_info)

    def save_dist(self, dist_data):
        """
        保存策略逻辑过程记录=》 csv文件按
        :param dist_data:
        :return:
        """
        if self.backtesting:
            save_path = self.cta_engine.get_logs_path()
        else:
            save_path = self.cta_engine.get_data_path()
        try:
            if 'margin' not in dist_data:
                dist_data.update({'margin': dist_data.get('price', 0) * dist_data.get('volume',
                                                                                      0) * self.cta_engine.get_margin_rate(
                    dist_data.get('symbol', self.vt_symbol))})
            if 'datetime' not in dist_data:
                dist_data.update({'datetime': self.cur_datetime})
            if self.position and 'long_pos' not in dist_data:
                dist_data.update({'long_pos': self.position.long_pos})
            if self.position and 'short_pos' not in dist_data:
                dist_data.update({'short_pos': self.position.short_pos})

            file_name = os.path.abspath(os.path.join(save_path, f'{self.strategy_name}_dist.csv'))
            append_data(file_name=file_name, dict_data=dist_data, field_names=self.dist_fieldnames)
        except Exception as ex:
            self.write_error(u'save_dist 异常:{} {}'.format(str(ex), traceback.format_exc()))

    def on_order(self, order: OrderData):
        """
        报单更新
        """
        if order.vt_orderid in self.active_orders:
            if order.volume == order.traded and order.status in [Status.ALLTRADED]:
                self.on_order_all_traded(order)

                # 更新仓位信息
                if order.offset == Offset.OPEN:
                    self.position.open_pos(order.direction, volume=order.volume)
                
                if order.offset != Offset.OPEN:
                    self.position.close_pos(order.direction, volume=order.volume)

            elif order.offset == Offset.OPEN and order.status in [Status.CANCELLED]:
                # 开仓委托单被撤销
                self.on_order_open_canceled(order)

            elif order.offset != Offset.OPEN and order.status in [Status.CANCELLED]:
                # 平仓委托单被撤销
                self.on_order_close_canceled(order)

            elif order.status == Status.REJECTED:
                if order.offset == Offset.OPEN:
                    self.write_error(u'{}委托单开{}被拒，price:{},total:{},traded:{}，status:{}'
                                        .format(order.vt_symbol, order.direction, order.price, order.volume,
                                                order.traded, order.status))
                    self.on_order_open_canceled(order)
                else:
                    self.write_error(u'OnOrder({})委托单平{}被拒，price:{},total:{},traded:{}，status:{}'
                                        .format(order.vt_symbol, order.direction, order.price, order.volume,
                                                order.traded, order.status))
                    self.on_order_close_canceled(order)
            else:
                self.write_log(u'委托单未完成,total:{},traded:{},tradeStatus:{}'
                                .format(order.volume, order.traded, order.status))
        else:
            self.write_error(u'委托单{}不在策略的未完成订单列表中:{}'.format(order.vt_orderid, self.active_orders))
    
    def on_order_all_traded(self, order: OrderData):
        """
        订单全部成交
        :param order:
        :return:
        """
        self.write_log(u'{},委托单:{}全部完成'.format(order.time, order.vt_orderid))
        order_info = self.active_orders[order.vt_orderid]

        # 通过vt_orderid，找到对应的网格
        grid:CtaGrid = order_info.get('grid', None)
        if grid is not None:
            # 移除当前委托单
            if order.vt_orderid in grid.order_ids:
                grid.order_ids.remove(order.vt_orderid)

            # 网格的所有委托单已经执行完毕
            if len(grid.order_ids) == 0:
                grid.order_status = False
                grid.traded_volume = 0

                if order.offset == Offset.OPEN:
                    # 开仓完毕( buy, short)
                    grid.open_status = True
                    grid.open_time = self.cur_datetime
                    grid.open_fee = order.fee
                    self.write_log(f'{grid.direction.value}单已开仓完毕,order_price:{order.price}'
                                   + f',volume:{order.volume}')
                else:
                    # 平仓完毕（cover， sell）
                    grid.open_status = False
                    grid.close_status = True
                    grid.close_fee = order.fee
                    self.save_trade(order, grid)

                    if grid.volume < order.traded:
                        self.write_log(f'网格平仓数量{grid.volume}，小于委托单成交数量:{order.volume}，修正为:{order.volume}')
                        grid.volume = order.traded

                    self.write_log(f'{grid.direction.value}单已平仓完毕,order_price:{order.price}'
                                   + f',volume:{order.volume}')

                    self.gt.remove_grids_by_ids(direction=grid.direction, ids=[grid.id])

                # 网格的所有委托单部分执行完毕
            else:
                old_traded_volume = grid.traded_volume
                grid.traded_volume += order.volume
                grid.traded_volume = round(grid.traded_volume, 7)
                grid.close_fee = order.fee
                self.save_trade(order, grid)

                self.write_log(f'{grid.direction.value}单部分{order.offset}仓，'
                               + f'网格volume:{grid.volume}, traded_volume:{old_traded_volume}=>{grid.traded_volume}')

                self.write_log(f'剩余委托单号:{grid.order_ids}')

            self.gt.save()

    def on_order_open_canceled(self, order: OrderData):
        """
        委托开仓单撤销
        :param order:
        :return:
        """
        self.write_log(u'委托开仓单撤销:{}'.format(order.__dict__))

        if order.vt_orderid not in self.active_orders:
            self.write_error(u'{}不在未完成的委托单中{}。'.format(order.vt_orderid, self.active_orders))
            return

        # 直接更新“未完成委托单”，更新volume,retry次数
        old_order = self.active_orders[order.vt_orderid]
        self.write_log(u'{} 委托信息:{}'.format(order.vt_orderid, old_order))
        old_order['traded'] = order.traded

        grid:CtaGrid = old_order.get('grid', None)

        pre_status = old_order.get('status', Status.NOTTRADED)
        if pre_status == Status.CANCELLED:
            self.write_log(f'当前状态已经是{Status.CANCELLED}，不做调整处理')
            return

        old_order.update({'status': Status.CANCELLED})
        self.write_log(u'委托单状态:{}=>{}'.format(pre_status, old_order.get('status')))
        if grid:
            if order.vt_orderid in grid.order_ids:
                grid.order_ids.remove(order.vt_orderid)
            if order.traded > 0:
                pre_traded_volume = grid.traded_volume
                grid.traded_volume = round(grid.traded_volume + order.traded, 7)
                self.write_log(f'撤单中部分开仓:{order.traded} + 原已成交:{pre_traded_volume}  => {grid.traded_volume}')
            if len(grid.order_ids) == 0:
                grid.order_status = False
                if grid.traded_volume > 0:
                    pre_volume = grid.volume
                    grid.volume = grid.traded_volume
                    grid.traded_volume = 0
                    grid.open_status = True
                    self.write_log(f'开仓完成，grid.volume {pre_volume} => {grid.volume}')

            self.gt.save()
        self.active_orders.update({order.vt_orderid: old_order})

    def on_order_close_canceled(self, order: OrderData):
        """
        委托平仓单撤销
        """
        self.write_log(u'委托平仓单撤销:{}'.format(order.__dict__))

        if order.vt_orderid not in self.active_orders:
            self.write_error(u'{}不在未完成的委托单中:{}。'.format(order.vt_orderid, self.active_orders))
            return

        # 直接更新“未完成委托单”，更新volume,Retry次数
        old_order = self.active_orders[order.vt_orderid]
        self.write_log(u'{} 订单信息:{}'.format(order.vt_orderid, old_order))
        old_order['traded'] = order.traded

        grid:CtaGrid = old_order.get('grid', None)
        pre_status = old_order.get('status', Status.NOTTRADED)
        if pre_status == Status.CANCELLED:
            self.write_log(f'当前状态已经是{Status.CANCELLED}，不做调整处理')
            return

        old_order.update({'status': Status.CANCELLED})
        self.write_log(u'委托单状态:{}=>{}'.format(pre_status, old_order.get('status')))
        if grid:
            if order.vt_orderid in grid.order_ids:
                grid.order_ids.remove(order.vt_orderid)
            if order.traded > 0:
                pre_traded_volume = grid.traded_volume
                grid.traded_volume = round(grid.traded_volume + order.traded, 7)
                self.write_log(f'撤单中部分平仓成交:{order.traded} + 原已成交:{pre_traded_volume}  => {grid.traded_volume}')
            if len(grid.order_ids) == 0:
                grid.order_status = False
                if grid.traded_volume > 0:
                    pre_volume = grid.volume
                    grid.volume = round(grid.volume - grid.traded_volume, 7)
                    grid.traded_volume = 0
                    if grid.volume <= 0:
                        grid.volume = 0
                        grid.open_status = False
                        self.write_log(f'强制全部平仓完成')
                    else:
                        self.write_log(f'平仓委托中，撤单完成，部分成交，减少持仓grid.volume {pre_volume} => {grid.volume}')

            self.gt.save()
        self.active_orders.update({order.vt_orderid: old_order})
        pass

    def grid_short(self, grid: CtaGrid, short_price, lock: bool = False, order_type: OrderType = OrderType.LIMIT):
        """
        事务开空仓
        :return:
        """
        vt_orderids = self.short(vt_symbol=self.vt_symbol,
                                 price=short_price,
                                 volume=grid.volume,
                                 order_type=order_type,
                                 order_time=self.cur_datetime,
                                 grid=grid,
                                 lock=lock)
        if len(vt_orderids) > 0:
            self.gt.save()
            return True
        else:
            self.write_error(u'创建{}事务空单,委托失败,开仓价：{}，数量：{}，止盈价:{}'
                             .format(grid.type, grid.open_price, grid.volume, grid.close_price))
            return False

    def grid_buy(self, grid: CtaGrid, buy_price, lock: bool = False, order_type: OrderType = OrderType.LIMIT):
        """
        事务开多仓
        :return:
        """
        vt_orderids = self.buy(vt_symbol=self.vt_symbol,
                               price=buy_price,
                               volume=grid.volume,
                               order_type=order_type,
                               order_time=self.cur_datetime,
                               grid=grid,
                               lock=lock)
        if len(vt_orderids) > 0:
            self.gt.save()
            return True
        else:
            self.write_error(u'创建{}事务多单,委托失败，开仓价：{}，数量：{}，止盈价:{}'
                             .format(grid.type, grid.open_price, grid.volume, grid.close_price))
            return False


    def grid_sell(self, grid: CtaGrid, sell_price, lock: bool = False, order_type: OrderType = OrderType.LIMIT):
        """
        事务平多单仓位
        1.来源自止损止盈平仓
        :param 平仓网格
        :return:
        """
        # 发出平多委托
        if grid.traded_volume > 0:
            grid.volume -= grid.traded_volume
            grid.volume = round(grid.volume, 7)
            grid.traded_volume = 0

        vt_orderids = self.sell(
            vt_symbol=self.vt_symbol,
            price=sell_price,
            volume=grid.volume,
            order_type=order_type,
            order_time=self.cur_datetime,
            grid=grid,
            lock=lock)
        if len(vt_orderids) == 0:
            self.write_error(u'多单平仓委托失败')
            return False
        else:
            self.write_log(u'多单平仓委托成功，编号:{}'.format(vt_orderids))

            return True

    def grid_cover(self, grid: CtaGrid, cover_price, lock: bool = False, order_type: OrderType = OrderType.LIMIT):
        """
        事务平空单仓位
        1.来源自止损止盈平仓
        :param 平仓网格
        :return:
        """
        # 发出cover委托
        if grid.traded_volume > 0:
            grid.volume -= grid.traded_volume
            grid.volume = round(grid.volume, 7)
            grid.traded_volume = 0

        vt_orderids = self.cover(
            price=cover_price,
            vt_symbol=self.vt_symbol,
            volume=grid.volume,
            order_type=order_type,
            order_time=self.cur_datetime,
            grid=grid,
            lock=lock)

        if len(vt_orderids) == 0:
            self.write_error(u'空单平仓委托失败')
            return False
        else:
            self.write_log(u'空单平仓委托成功，编号:{}'.format(vt_orderids))
            return True

    def save_trade(self, order: OrderData, grid: CtaGrid):
        """
        将交易费用保存于excel
        """
        if self.get_engine_type() == EngineType.BACKTESTING:
            return

        trade_save_path = get_folder_path('data')
        xlsx_name = f"{self.strategy_name}_trade.xlsx"
        xlsx_file = str(trade_save_path.joinpath(xlsx_name))

        if os.path.exists(xlsx_file):
            workbook = load_workbook(xlsx_file)
            my_sheet = workbook.get_sheet_by_name("Mysheet")
            row = my_sheet.max_row
            row += 1
        else:
            workbook = Workbook()
            my_sheet = workbook.create_sheet("Mysheet", index=0)
            
            row = 1
            head = ['品种', '开仓时间', '平仓时间', '方向', '数量', '开仓价格', '平仓价格', '开仓手续费', '平仓手续费', '平仓收益' , '净收入']
            for i, item in enumerate(head):
                my_sheet.cell(row, i + 1, item)
            my_sheet.column_dimensions["A"].width = 20
            my_sheet.column_dimensions["B"].width = 20
            my_sheet.column_dimensions["C"].width = 20
            row += 1

        my_sheet.cell(row, 1, grid.vt_symbol)
        my_sheet.cell(row, 2, grid.open_time)
        my_sheet.cell(row, 3, order.datetime)
        my_sheet.cell(row, 4, grid.direction.value if isinstance(grid.direction, Direction) else '')
        my_sheet.cell(row, 5, grid.volume)
        my_sheet.cell(row, 6, grid.open_price)
        my_sheet.cell(row, 7, order.price)
        my_sheet.cell(row, 8, grid.open_fee)
        my_sheet.cell(row, 9, grid.close_fee)
        
        my_sheet.cell(row, 10, order.netpnl)
        my_sheet.cell(row, 11, order.netpnl - grid.open_fee - grid.close_fee)
        
        workbook.save(xlsx_file)


